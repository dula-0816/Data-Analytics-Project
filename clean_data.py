import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

# Load dataset
df = pd.read_excel("Dataset for Data Analytics.xlsx")

#Initial audit of the dataset
print("--- INITIAL AUDIT ---")
print("Missing values per column before cleaning:\n", df.isnull().sum())
print("\nNumber of duplicate rows:", df.duplicated().sum())
print("Duplicate OrderIDs:", df["OrderID"].duplicated().sum())
print("---------------------\n")

# 1. Check missing values
print("Missing values per column:\n", df.isnull().sum())

# 2. Handle missing values
df["CouponCode"] = df["CouponCode"].fillna("N/A")

# 3. Remove duplicate rows
print("Number of duplicate rows:", df.duplicated().sum())
df = df.drop_duplicates()

# 4. Remove duplicate OrderIDs
print("Duplicate OrderIDs:", df["OrderID"].duplicated().sum())
df = df.drop_duplicates(subset=["OrderID"])

# 5. Fix date format
df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%Y-%m-%d")

# 6. Final verification check
print("--- FINAL VERIFICATION ---")
print("Missing values remaining:\n", df.isnull().sum())
print("--------------------------")

# 7. Save cleaned dataset as NEW Excel file
output_file = "Cleaned_Dataset_Data_Analytics.xlsx"

df.to_excel(output_file, index=False)

# Load workbook
workbook = load_workbook(output_file)

# Select active sheet
sheet = workbook.active

# Bold the header row
for cell in sheet[1]:
    cell.font = Font(bold=True)

# Center align all cells
for row in sheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Save workbook
workbook.save(output_file)

print("\nData cleaning completed and saved successfully!")